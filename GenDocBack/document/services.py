"""
Service de génération de documents.
Gère la création synchrone des fichiers Word à partir des templates.
"""

import os
import tempfile
from pathlib import Path
from django.utils import timezone
from django.core.files.base import ContentFile
from .models import Document
from templates.models import TemplateVersion


class DocumentService:
    """
    Service pour générer les documents de manière synchrone.
    """

    @staticmethod
    def generate(template_version: TemplateVersion, input_data: dict) -> tuple:
        """
        Génère un document synchroniquement.
        
        Args:
            template_version: L'objet TemplateVersion contenant le template
            input_data: Dict avec les données à injecter dans le template
            
        Returns:
            tuple: (fichier_genere, nom_fichier)
            
        Raises:
            ValueError: Si la génération échoue
        """
        try:
            # Récupère le chemin du fichier template
            template_path = template_version.source_file.path
            
            if not os.path.exists(template_path):
                raise FileNotFoundError(f"Template non trouvé : {template_path}")
            
            # Détecte le format du template (DOCX pour Word)
            file_ext = os.path.splitext(template_path)[1].lower()
            
            if file_ext == '.docx':
                return DocumentService._generate_docx(template_path, input_data)
            elif file_ext == '.xlsx':
                return DocumentService._generate_xlsx(template_path, input_data)
            else:
                raise ValueError(f"Format de template non supporté : {file_ext}")
                
        except Exception as e:
            raise ValueError(f"Erreur lors de la génération : {str(e)}")

    @staticmethod
    def _generate_docx(template_path: str, input_data: dict) -> tuple:
        """
        Génère un fichier DOCX en injectant les variables du template.
        
        Utilise python-docx ou docxtpl selon la méthode.
        """
        try:
            from docxtpl import DocxTemplate
        except ImportError:
            raise ImportError("Installez 'python-docx' ou 'docxtpl' pour générer des fichiers DOCX")
        
        try:
            # Charge le template
            doc = DocxTemplate(template_path)
            
            # Injecte les données
            doc.render(input_data)
            
            # Crée un fichier temporaire
            with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as tmp:
                doc.save(tmp.name)
                tmp_path = tmp.name
            
            # Lit le contenu du fichier généré
            with open(tmp_path, 'rb') as f:
                file_content = f.read()
            
            # Nettoie le fichier temporaire
            os.unlink(tmp_path)
            
            # Génère un nom de fichier
            filename = DocumentService._generate_filename(input_data, '.docx')
            
            return file_content, filename
            
        except Exception as e:
            raise ValueError(f"Erreur DOCX : {str(e)}")

    @staticmethod
    def _generate_xlsx(template_path: str, input_data: dict) -> tuple:
        """
        Génère un fichier XLSX en injectant les variables du template.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("Installez 'openpyxl' pour générer des fichiers XLSX")
        
        try:
            # Charge le template
            workbook = load_workbook(template_path)
            
            # Injecte les données dans les cellules
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            # Remplace les variables {{var}}
                            for key, value in input_data.items():
                                cell.value = str(cell.value).replace(f"{{{{{key}}}}}", str(value))
            
            # Crée un fichier temporaire
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                workbook.save(tmp.name)
                tmp_path = tmp.name
            
            # Lit le contenu du fichier généré
            with open(tmp_path, 'rb') as f:
                file_content = f.read()
            
            # Nettoie le fichier temporaire
            os.unlink(tmp_path)
            
            # Génère un nom de fichier
            filename = DocumentService._generate_filename(input_data, '.xlsx')
            
            return file_content, filename
            
        except Exception as e:
            raise ValueError(f"Erreur XLSX : {str(e)}")

    @staticmethod
    def _generate_filename(input_data: dict, extension: str) -> str:
        """
        Génère un nom de fichier à partir des données d'entrée.
        """
        # Utilise le champ 'nom' ou 'name' s'il existe, sinon utilise le timestamp
        base_name = input_data.get('nom') or input_data.get('name') or 'document'
        
        # Nettoie le nom de fichier (enlève les caractères spéciaux)
        base_name = ''.join(c for c in base_name if c.isalnum() or c in '-_ ')
        
        # Ajoute le timestamp pour l'unicité
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        return f"{base_name}_{timestamp}{extension}"
